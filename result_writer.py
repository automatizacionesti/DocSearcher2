from typing import List, Dict
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.hyperlink import Hyperlink
import time

class ResultWriter:
    def __init__(self):
        # Usar ruta absoluta para el directorio de resultados
        self.results_dir = os.path.abspath('resultados')

    def write_results(self, results):
        """Escribe los resultados en un archivo Excel."""
        try:
            # Validar que results no esté vacío
            if not results:
                print("No hay resultados para escribir en el archivo Excel")
                return None
                
            # Validar estructura de resultados
            print(f"Estructura de resultados recibida: {type(results)}")
            print(f"Número de resultados: {len(results)}")
            print(f"Primer resultado: {results[0] if results else None}")
            
            # Crear directorio si no existe
            try:
                os.makedirs(self.results_dir, exist_ok=True)
                # En sistemas Unix, dar permisos completos
                if os.name != 'nt':
                    os.chmod(self.results_dir, 0o777)
            except Exception as e:
                print(f"Error creando directorio: {str(e)}")
                raise
            
            # Generar nombre de archivo único
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'resultados_{timestamp}.xlsx'
            filepath = os.path.join(self.results_dir, filename)
            
            print(f"Intentando escribir en: {filepath}")
            
            # Crear nuevo libro de Excel
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = 'Resultados'
            
            # Configurar estilos para encabezados
            header_style = {
                'font': Font(bold=True, color='FFFFFF'),
                'fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center')
            }
            
            # Definir encabezados
            headers = [
                'Nombre de archivo',
                'Ruta del archivo',
                'Palabra encontrada',
                'Contexto'
            ]
            
            # Escribir encabezados con estilo
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col, value=header)
                cell.font = header_style['font']
                cell.fill = header_style['fill']
                cell.alignment = header_style['alignment']
            
            # Escribir datos
            row = 2
            total_rows_written = 0
            try:
                for result in results:
                    if not isinstance(result, dict):
                        print(f"Saltando resultado no válido: {type(result)}")
                        continue
                        
                    file_name = result.get('file_name', '')
                    findings = result.get('findings', [])
                    print(f"Procesando archivo: {file_name} con {len(findings)} hallazgos")
                    
                    for finding in findings:
                        if not isinstance(finding, dict):
                            print(f"Saltando hallazgo no válido: {type(finding)}")
                            continue
                            
                        try:
                            # Escribir fila de datos
                            sheet.cell(row=row, column=1, value=str(file_name))
                            sheet.cell(row=row, column=2, value=str(finding.get('path', '')))
                            sheet.cell(row=row, column=3, value=str(finding.get('palabra', '')))
                            sheet.cell(row=row, column=4, value=str(finding.get('contexto', '')))
                            row += 1
                            total_rows_written += 1
                        except Exception as e:
                            print(f"Error escribiendo fila {row}: {str(e)}")
                            continue
                
                print(f"Total de filas escritas: {total_rows_written}")
                if total_rows_written == 0:
                    raise Exception("No se escribieron datos en el archivo Excel")
                
                # Ajustar ancho de columnas
                for column_cells in sheet.columns:
                    length = max(len(str(cell.value)) for cell in column_cells)
                    sheet.column_dimensions[column_cells[0].column_letter].width = length + 2
                
                # Guardar archivo
                try:
                    workbook.save(filepath)
                    time.sleep(1)  # Esperar a que el archivo se escriba
                    print(f"Archivo guardado con éxito en: {filepath}")
                    print(f"Tamaño del archivo: {os.path.getsize(filepath)} bytes")
                    
                    # Verificar permisos del archivo
                    if os.name != 'nt':
                        os.chmod(filepath, 0o666)
                except Exception as e:
                    print(f"Error al guardar el archivo: {str(e)}")
                    raise
            finally:
                workbook.close()
            
            # Verificar que se guardó correctamente
            if os.path.exists(filepath) and os.path.getsize(filepath) > 0:
                time.sleep(0.5)  # Esperar a que el sistema de archivos se actualice
                return filename
            else:
                raise Exception("Error al guardar el archivo Excel")
                
        except Exception as e:
            print(f"Error al crear archivo Excel: {str(e)}")
            if 'filepath' in locals() and os.path.exists(filepath):
                try:
                    os.remove(filepath)
                except:
                    pass
            raise 